# import openpyxl
# import os

# # تحديد نطاق الروابط
# base_url = "https://ufogrendizer.tv/watch/269/"
# start_episode = 2469
# end_episode = 2510

# # إنشاء قائمة لتخزين الروابط
# urls = [f"{base_url}{episode}" for episode in range(start_episode, end_episode + 1)]

# # إنشاء ملف Excel جديد
# workbook = openpyxl.Workbook()
# sheet = workbook.active
# sheet.title = "Episode Links"

# # كتابة الروابط في الملف
# sheet.append(["Episode Number", "Link"])  # كتابة العناوين

# for episode, url in enumerate(urls, start=start_episode):
#     sheet.append([episode, url])

# # تحديد المسار على سطح المكتب
# desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
# file_path = os.path.join(desktop_path, "episode_links.xlsx")

# # حفظ الملف
# workbook.save(file_path)
# print(f"File saved to: {file_path}")


import requests
from bs4 import BeautifulSoup
import openpyxl
import csv
import os
import time  # استيراد مكتبة time لاستخدام وظيفة sleep

# تحديد مسار ملف Excel على سطح المكتب
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
excel_file_path = os.path.join(desktop_path, "episode_links.xlsx")

# قراءة الروابط من ملف Excel
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

# استخراج الروابط من الملف
links = []
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
    links.append(row[0].value)

# استخراج روابط الفيديو من كل صفحة
video_links = []

for url in links:
    try:
        response = requests.get(url)
        response.raise_for_status()  # التحقق من استجابة الطلب
        soup = BeautifulSoup(response.content, "html.parser")
        
        # البحث عن أول علامة فيديو
        video_tag = soup.find("video", {"src": True})
        
        if video_tag:
            video_src = video_tag['src']
            video_links.append(video_src)
            print(f"Found video link for {url}: {video_src}")
        else:
            print(f"Video link not found for {url}.")
            video_links.append("Not found")
            
    except requests.exceptions.RequestException as e:
        print(f"Error fetching {url}: {e}")
        video_links.append("Error")
    
    # فاصل زمني (3 ثوانٍ) بين كل طلب
    time.sleep(1)

# حفظ روابط الفيديو في ملف CSV على سطح المكتب
csv_file_path = os.path.join(desktop_path, "video_links.csv")

with open(csv_file_path, mode='w', newline='', encoding='utf-8') as file:
    writer = csv.writer(file)
    writer.writerow(["Episode Link", "Video Link"])
    for episode_link, video_link in zip(links, video_links):
        writer.writerow([episode_link, video_link])

print(f"Video links saved to: {csv_file_path}")
